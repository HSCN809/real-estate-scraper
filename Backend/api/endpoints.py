from fastapi import APIRouter, HTTPException, BackgroundTasks, Depends
from fastapi.responses import FileResponse
from api.schemas import ScrapeRequest, ScrapeResponse
from core.driver_manager import DriverManager
from core.config import get_emlakjet_config, get_hepsiemlak_config
from api.status import task_status
from sqlalchemy.orm import Session
from database.connection import get_db
from database import crud
from database.models import Listing, Location, ScrapeSession
import logging
from typing import Optional, List
from datetime import datetime
import os
import io

# Import scrapers (will need refactoring to import cleanly)
# We will do dynamic imports or ensure the refactor makes them importable
# from scrapers.emlakjet.main import EmlakJetScraper
# from scrapers.hepsiemlak.main import HepsiemlakScraper

router = APIRouter()
logger = logging.getLogger("api")

# Category name mappings for display
CATEGORY_NAMES = {
    "konut": "Konut",
    "arsa": "Arsa",
    "isyeri": "İşyeri",
    "devremulk": "Devremülk",
    "turistik_isletme": "Turistik İşletme",
    "turistik_tesis": "Turistik Tesis",
    "kat_karsiligi_arsa": "Kat Karşılığı Arsa",
    "devren_isyeri": "Devren İşyeri",
    "gunluk_kiralik": "Günlük Kiralık"
}

@router.get("/config/categories")
async def get_categories():
    """Get all platform categories from config"""
    emlakjet = get_emlakjet_config()
    hepsiemlak = get_hepsiemlak_config()
    
    def format_categories(categories_dict):
        return [
            {"id": k, "name": CATEGORY_NAMES.get(k, k.replace("_", " ").title())}
            for k in categories_dict.keys()
        ]
    
    return {
        "emlakjet": {
            "satilik": format_categories(emlakjet.categories["satilik"]),
            "kiralik": format_categories(emlakjet.categories["kiralik"])
        },
        "hepsiemlak": {
            "satilik": format_categories(hepsiemlak.categories["satilik"]),
            "kiralik": format_categories(hepsiemlak.categories["kiralik"])
        }
    }

@router.get("/config/subtypes")
async def get_subtypes(listing_type: str, category: str):
    """
    Belirli bir kategori için alt tipleri JSON dosyasından oku.
    JSON dosyası yoksa boş liste döner.
    Örnek: /config/subtypes?listing_type=kiralik&category=arsa
    """
    from scrapers.hepsiemlak.subtype_fetcher import fetch_subtypes, SUBCATEGORIES_JSON_PATH
    
    # JSON dosyasından oku
    subtypes = fetch_subtypes(listing_type, category)
    
    if not SUBCATEGORIES_JSON_PATH.exists():
        return {
            "subtypes": [],
            "error": "Subcategories JSON dosyası bulunamadı. 'python -m scrapers.hepsiemlak.subtype_fetcher' çalıştırın."
        }
    
    return {"subtypes": subtypes, "cached": True}

def run_emlakjet_task(request: ScrapeRequest):
    from scrapers.emlakjet.main import EmlakJetScraper
    
    manager = DriverManager()
    try:
        driver = manager.start()
        config = get_emlakjet_config()
        
        # Reset and start status
        task_status.reset()
        task_status.set_running(True)
        task_status.update("EmlakJet scraper başlatılıyor...", progress=0)

        def progress_callback(message, current=0, total=0, progress=0):
            task_status.update(message=message, current=current, total=total, progress=progress)
        
        # Construct base URL based on inputs
        # Logic adapted from original main.py
        category_path = config.categories[request.listing_type].get(request.category, '')
        base_url = config.base_url + category_path
        
        # Pass parameters to scraper (Refactor required on Scraper class)
        # Assuming we will refactor the scraper to accept direct options
        scraper = EmlakJetScraper(
            driver=driver,
            base_url=base_url,
            category=request.category,
            listing_type=request.listing_type  # Klasör yapısı için
            # We will add a new init param or method to set simple mode
            # simple_mode=True
        )
        
        # For now, we need to handle the interactive parts or bypass them.
        # This part assumes the REFACTOR step has been done to add `start_scraping_api` or similar.
        if hasattr(scraper, 'start_scraping_api'):
             scraper.start_scraping_api(
                 cities=request.cities,
                 districts=request.districts,
                 max_pages=request.max_pages,
                 progress_callback=progress_callback
             )
        else:
            logger.warning("Scraper api method not found (Refactor needed)")
            
    except Exception as e:
        logger.error(f"EmlakJet task error: {e}")
    finally:
        task_status.set_running(False)
        task_status.update("İşlem tamamlandı", progress=100)
        manager.stop()

def run_hepsiemlak_task(request: ScrapeRequest):
    from scrapers.hepsiemlak.main import HepsiemlakScraper
    from core.failed_pages_tracker import failed_pages_tracker
    from database.connection import get_db_session
    from database import crud

    manager = DriverManager()
    db = None
    scrape_session = None

    try:
        driver = manager.start()
        db = get_db_session()

        # Reset trackers
        task_status.reset()
        failed_pages_tracker.reset()

        task_status.set_running(True)
        task_status.update("HepsiEmlak scraper başlatılıyor...", progress=0)

        # Alt kategori adını çıkar
        alt_kategori = None
        if request.subtype_path:
            parts = request.subtype_path.strip('/').split('/')
            if len(parts) >= 2:
                alt_kategori = parts[-1].replace('-', '_')

        # Veritabanında scrape session oluştur
        scrape_session = crud.create_scrape_session(
            db,
            platform="hepsiemlak",
            kategori=request.category,
            ilan_tipi=request.listing_type,
            alt_kategori=alt_kategori,
            target_cities=request.cities,
            target_districts=request.districts
        )
        db.commit()
        logger.info(f"ScrapeSession created: ID={scrape_session.id}")

        def progress_callback(message, current=0, total=0, progress=0):
            task_status.update(message=message, current=current, total=total, progress=progress)

        scraper = HepsiemlakScraper(
            driver=driver,
            listing_type=request.listing_type,
            category=request.category,
            subtype_path=request.subtype_path,
            selected_cities=request.cities,
            selected_districts=request.districts
        )

        # Scraper'a DB session'ı ve session_id'yi ekle
        scraper.db = db
        scraper.scrape_session_id = scrape_session.id

        print(f"DEBUG API: listing_type={request.listing_type}, category={request.category}, subtype_path={request.subtype_path}, cities={request.cities}")

        if hasattr(scraper, 'start_scraping_api'):
            scraper.start_scraping_api(max_pages=request.max_pages, progress_callback=progress_callback)
        else:
            logger.warning("Scraper api method not found (Refactor needed)")

        # Scrape tamamlandıktan sonra session'ı güncelle
        if scrape_session:
            # Scraper'dan toplam ilan sayısını al
            total_listings = getattr(scraper, 'total_scraped_count', 0)
            new_listings = getattr(scraper, 'new_listings_count', 0)
            duplicate_listings = getattr(scraper, 'duplicate_count', 0)

            crud.update_scrape_session(
                db,
                scrape_session.id,
                total_listings=total_listings,
                new_listings=new_listings,
                duplicate_listings=duplicate_listings
            )
            crud.complete_scrape_session(db, scrape_session.id, status="completed")
            db.commit()
            logger.info(f"ScrapeSession completed: ID={scrape_session.id}, total={total_listings}")

    except Exception as e:
        logger.error(f"HepsiEmlak task error: {e}")
        if scrape_session and db:
            crud.complete_scrape_session(db, scrape_session.id, status="failed", error_message=str(e))
            db.commit()
    finally:
        task_status.set_running(False)
        task_status.update("İşlem tamamlandı", progress=100)
        if db:
            db.close()
        manager.stop()

@router.post("/scrape/emlakjet", response_model=ScrapeResponse)
async def scrape_emlakjet(request: ScrapeRequest, background_tasks: BackgroundTasks):
    # Status'u HEMEN ayarla
    task_status.reset()
    task_status.set_running(True)
    task_status.update("EmlakJet taraması başlatılıyor...", progress=0)
    
    background_tasks.add_task(run_emlakjet_task, request)
    return ScrapeResponse(
        status="accepted",
        message="EmlakJet scraping task started in background",
        data_count=0,
        output_files=[]
    )

@router.post("/scrape/hepsiemlak", response_model=ScrapeResponse)
async def scrape_hepsiemlak(request: ScrapeRequest, background_tasks: BackgroundTasks):
    # Validate cities
    if not request.cities or len(request.cities) == 0:
        raise HTTPException(status_code=400, detail="En az bir şehir seçmelisiniz")

    # Validate districts if provided
    if request.districts:
        for city, districts in request.districts.items():
            if city not in request.cities:
                raise HTTPException(
                    status_code=400,
                    detail=f"İlçe seçilen şehir ({city}) şehir listesinde yok"
                )

    # Status'u HEMEN ayarla - frontend ilk sorguladığında hazır olsun
    task_status.reset()
    task_status.set_running(True)
    task_status.update("HepsiEmlak taraması başlatılıyor...", progress=0)

    background_tasks.add_task(run_hepsiemlak_task, request)
    return ScrapeResponse(
        status="accepted",
        message="HepsiEmlak scraping task started in background",
        data_count=0,
        output_files=[]
    )

@router.post("/stop")
async def stop_scraping():
    """Aktif tarama işlemini durdur ve mevcut verileri kaydet"""
    if task_status.is_running:
        task_status.request_stop()
        return {"status": "stopping", "message": "Durdurma isteği gönderildi. Mevcut veriler kaydediliyor..."}
    else:
        return {"status": "idle", "message": "Aktif bir tarama işlemi yok."}

@router.get("/analytics/prices")
async def get_price_analytics(
    platform: str = None,
    category: str = None,
    listing_type: str = None,
    db: Session = Depends(get_db)
):
    """Veritabanından fiyat verilerini çek - grafikler için (filtrelenebilir)"""
    # Platform/category/listing_type mapping for display names
    platform_map = {"hepsiemlak": "HepsiEmlak", "emlakjet": "Emlakjet"}
    category_map = {"konut": "Konut", "arsa": "Arsa", "isyeri": "İşyeri", "devremulk": "Devremülk"}
    listing_type_map = {"satilik": "Satılık", "kiralik": "Kiralık"}

    # Convert display names back to db values for filtering
    db_platform = None
    db_category = None
    db_listing_type = None

    if platform and platform != "all":
        db_platform = platform.lower().replace("hepsiemlak", "hepsiemlak").replace("emlakjet", "emlakjet")
        if platform == "HepsiEmlak":
            db_platform = "hepsiemlak"
        elif platform == "Emlakjet":
            db_platform = "emlakjet"

    if category and category != "all":
        for k, v in category_map.items():
            if v == category:
                db_category = k
                break
        if not db_category:
            db_category = category.lower()

    if listing_type and listing_type != "all":
        for k, v in listing_type_map.items():
            if v == listing_type:
                db_listing_type = k
                break
        if not db_listing_type:
            db_listing_type = listing_type.lower()

    result = crud.get_price_analytics(
        db,
        platform=db_platform,
        kategori=db_category,
        ilan_tipi=db_listing_type
    )

    # Transform to display names
    prices = []
    for p in result["prices"]:
        prices.append({
            "city": p["city"],
            "platform": platform_map.get(p["platform"], p["platform"]),
            "category": category_map.get(p["category"], p["category"]),
            "listing_type": listing_type_map.get(p["listing_type"], p["listing_type"]),
            "price": p["price"]
        })

    return {"prices": prices, "summary": result["summary"]}

@router.get("/analytics/city/{city_name}")
async def get_city_analytics(
    city_name: str,
    platform: str = None,
    category: str = None,
    listing_type: str = None,
    subtype: str = None,
    db: Session = Depends(get_db)
):
    """Veritabanından belirli bir şehrin ilanlarını ve istatistiklerini döndür"""
    # Convert display names to db values
    db_platform = None
    db_category = None
    db_listing_type = None

    if platform and platform != "all":
        if platform == "HepsiEmlak":
            db_platform = "hepsiemlak"
        elif platform == "Emlakjet":
            db_platform = "emlakjet"
        else:
            db_platform = platform.lower()

    if category and category != "all":
        category_map = {"Konut": "konut", "Arsa": "arsa", "İşyeri": "isyeri", "Devremülk": "devremulk"}
        db_category = category_map.get(category, category.lower())

    if listing_type and listing_type != "all":
        listing_map = {"Satılık": "satilik", "Kiralık": "kiralik"}
        db_listing_type = listing_map.get(listing_type, listing_type.lower())

    return crud.get_city_analytics(
        db,
        city_name=city_name,
        platform=db_platform,
        kategori=db_category,
        ilan_tipi=db_listing_type
    )

@router.get("/analytics/file-stats/{filename}")
async def get_file_statistics(filename: str):
    """Dosya bazlı detaylı istatistikler - describe + fiyat aralıkları"""
    import os
    import re
    import statistics
    
    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
    
    output_dir = os.path.join(project_root, "outputs")
    if not os.path.exists(output_dir):
        output_dir = os.path.join(project_root, "Outputs")
    
    # Find the file
    file_path = None
    for root, dirs, files in os.walk(output_dir):
        if filename in files:
            file_path = os.path.join(root, filename)
            break
    
    if not file_path or not os.path.exists(file_path):
        return {"error": "Dosya bulunamadı", "stats": None}
    
    prices = []
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Find price column
        price_col_idx = None
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        
        for idx, cell in enumerate(header_row):
            val = str(cell.value).strip().lower() if cell.value else ""
            if val == "fiyat":
                price_col_idx = idx
                break
            elif "fiyat" in val and "metrekare" not in val:
                if price_col_idx is None:
                    price_col_idx = idx
        
        if price_col_idx is not None:
            for row in ws.iter_rows(min_row=2):
                cell = row[price_col_idx]
                if cell.value:
                    try:
                        val_str = str(cell.value).strip()
                        val_str = re.sub(r'\s+', '', val_str)
                        val_str = val_str.replace('TL', '').replace('₺', '')
                        
                        if ',' in val_str and '.' in val_str:
                            val_str = val_str.replace('.', '').replace(',', '.')
                        elif '.' in val_str and val_str.count('.') > 1:
                            val_str = val_str.replace('.', '')
                        elif '.' in val_str and len(val_str.split('.')[-1]) == 3:
                            val_str = val_str.replace('.', '')
                        elif ',' in val_str:
                            val_str = val_str.replace(',', '.')
                        
                        price = float(val_str)
                        if price > 0:
                            prices.append(price)
                    except:
                        pass
        
        wb.close()
    except Exception as e:
        logger.error(f"Error reading file stats: {e}")
        return {"error": str(e), "stats": None}
    
    if not prices:
        return {"error": "Fiyat verisi bulunamadı", "stats": None}
    
    # Sort prices for percentile calculations
    sorted_prices = sorted(prices)
    n = len(sorted_prices)
    
    # Calculate percentiles (quartiles)
    def percentile(data, p):
        k = (len(data) - 1) * p / 100
        f = int(k)
        c = f + 1 if f + 1 < len(data) else f
        return data[f] + (data[c] - data[f]) * (k - f) if c < len(data) else data[f]
    
    # Descriptive statistics (like pandas describe)
    describe_stats = {
        "count": n,
        "mean": round(statistics.mean(prices), 2),
        "std": round(statistics.stdev(prices), 2) if n > 1 else 0,
        "min": round(min(prices), 2),
        "q25": round(percentile(sorted_prices, 25), 2),
        "median": round(statistics.median(prices), 2),
        "q75": round(percentile(sorted_prices, 75), 2),
        "max": round(max(prices), 2)
    }
    
    # Dynamic price range distribution using quantile-based binning (like pandas qcut)
    # Create 5 bins based on data distribution
    num_bins = 5
    bin_edges = [percentile(sorted_prices, i * 100 / num_bins) for i in range(num_bins + 1)]
    
    # Ensure unique edges
    unique_edges = []
    for e in bin_edges:
        if not unique_edges or e > unique_edges[-1]:
            unique_edges.append(e)
    
    # If not enough unique edges, fall back to equal-width bins
    if len(unique_edges) < 3:
        min_p, max_p = min(prices), max(prices)
        bin_width = (max_p - min_p) / num_bins
        unique_edges = [min_p + i * bin_width for i in range(num_bins + 1)]
    
    # Count items in each bin
    price_ranges = []
    for i in range(len(unique_edges) - 1):
        low = unique_edges[i]
        high = unique_edges[i + 1]
        
        # Format the range label
        def format_price(p):
            if p >= 1000000:
                return f"{p/1000000:.1f}M"
            elif p >= 1000:
                return f"{p/1000:.0f}K"
            else:
                return f"{p:.0f}"
        
        label = f"{format_price(low)} - {format_price(high)}"
        
        # Count items in range
        if i == len(unique_edges) - 2:  # Last bin includes upper edge
            count = sum(1 for p in prices if low <= p <= high)
        else:
            count = sum(1 for p in prices if low <= p < high)
        
        price_ranges.append({
            "range": label,
            "count": count,
            "percentage": round(count / n * 100, 1)
        })
    
    return {
        "stats": describe_stats,
        "price_ranges": price_ranges,
        "total_listings": n
    }

@router.delete("/clear-results")
async def clear_results():
    """Outputs klasöründeki tüm sonuç dosyalarını sil"""
    import os
    import shutil
    
    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
    output_dir = os.path.join(project_root, "outputs")
    
    if not os.path.exists(output_dir):
        output_dir = os.path.join(project_root, "Outputs")
    
    if not os.path.exists(output_dir):
        return {"status": "error", "message": "Outputs klasörü bulunamadı", "deleted_count": 0}
    
    deleted_count = 0
    try:
        for item in os.listdir(output_dir):
            item_path = os.path.join(output_dir, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
                deleted_count += 1
            elif os.path.isdir(item_path):
                shutil.rmtree(item_path)
                deleted_count += 1
        
        return {"status": "success", "message": f"{deleted_count} dosya/klasör silindi", "deleted_count": deleted_count}
    except Exception as e:
        return {"status": "error", "message": str(e), "deleted_count": deleted_count}

@router.delete("/results/{filename:path}")
async def delete_result(filename: str):
    """Tek bir sonuç dosyasını sil"""
    import os
    import gc
    import time
    from urllib.parse import unquote
    
    # URL decode the filename (handles Turkish characters)
    decoded_filename = unquote(filename)
    logger.info(f"Delete request for: {decoded_filename}")
    
    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
    output_dir = os.path.join(project_root, "outputs")
    
    if not os.path.exists(output_dir):
        output_dir = os.path.join(project_root, "Outputs")
    
    # Find the file
    file_path = None
    for root, dirs, files in os.walk(output_dir):
        for f in files:
            if f == decoded_filename:
                file_path = os.path.join(root, f)
                break
        if file_path:
            break
    
    if not file_path or not os.path.exists(file_path):
        logger.warning(f"File not found: {decoded_filename}")
        raise HTTPException(status_code=404, detail=f"Dosya bulunamadı: {decoded_filename}")
    
    # Retry mechanism for Windows file locking
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Force garbage collection to release any file handles
            gc.collect()
            time.sleep(0.1)  # Small delay
            
            os.remove(file_path)
            logger.info(f"Deleted: {file_path}")
            return {"status": "success", "message": f"{decoded_filename} silindi"}
        except PermissionError as e:
            if attempt < max_retries - 1:
                logger.warning(f"Retry {attempt + 1}/{max_retries} for {decoded_filename}")
                gc.collect()
                time.sleep(0.5)  # Wait longer before retry
            else:
                logger.error(f"Delete failed after {max_retries} attempts: {e}")
                raise HTTPException(
                    status_code=500, 
                    detail="Dosya başka bir işlem tarafından kullanılıyor. Lütfen biraz bekleyip tekrar deneyin."
                )
        except Exception as e:
            logger.error(f"Delete error: {e}")
            raise HTTPException(status_code=500, detail=str(e))


@router.get("/results")
async def get_results(db: Session = Depends(get_db)):
    """
    Veritabanından sonuçları döndür.
    Şehir/platform/kategori bazında gruplanmış ilan özetleri.
    """
    return crud.get_results_for_frontend(db)

@router.get("/results/{filename}/preview")
async def get_result_preview(filename: str, limit: int = 20):
    """Dosyadan ilk N kaydı döndür"""
    import os
    from datetime import datetime
    
    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
    
    # Find the file
    output_dir = os.path.join(project_root, "outputs")
    if not os.path.exists(output_dir):
        output_dir = os.path.join(project_root, "Outputs")
    
    file_path = None
    for root, dirs, files in os.walk(output_dir):
        if filename in files:
            file_path = os.path.join(root, filename)
            break
    
    if not file_path or not os.path.exists(file_path):
        return {"error": "Dosya bulunamadı", "data": [], "total": 0}
    
    data = []
    total = 0
    
    try:
        if filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Get headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else "")
            
            # Get data rows
            total = ws.max_row - 1 if ws.max_row else 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(limit + 1, ws.max_row)), start=1):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        row_data[headers[col_idx]] = cell.value if cell.value else ""
                data.append(row_data)
            
            wb.close()
            
        elif filename.endswith('.json'):
            import json
            with open(file_path, 'r', encoding='utf-8') as f:
                all_data = json.load(f)
                if isinstance(all_data, list):
                    total = len(all_data)
                    data = all_data[:limit]
    except Exception as e:
        logger.error(f"Preview error: {e}")
        return {"error": str(e), "data": [], "total": 0}
    
    return {"data": data, "total": total, "showing": len(data)}

@router.get("/status")
async def get_status():
    return task_status.to_dict()

@router.get("/stats")
async def get_stats(db: Session = Depends(get_db)):
    """Veritabanından genel istatistikleri döndür"""
    return crud.get_stats_summary(db)

@router.get("/download/{filename}")
async def download_file(filename: str):
    """Dosyayı indir"""
    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))

    # Find the file
    output_dir = os.path.join(project_root, "outputs")
    if not os.path.exists(output_dir):
        output_dir = os.path.join(project_root, "Outputs")

    file_path = None
    for root, dirs, files in os.walk(output_dir):
        if filename in files:
            file_path = os.path.join(root, filename)
            break

    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Dosya bulunamadı")

    # Determine media type
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.endswith('.xlsx') else "application/json"

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type=media_type
    )


# ==================== NEW DB-BASED ENDPOINTS ====================

@router.get("/listings")
async def get_listings(
    platform: str = None,
    kategori: str = None,
    ilan_tipi: str = None,
    alt_kategori: str = None,
    city: str = None,
    district: str = None,
    min_price: float = None,
    max_price: float = None,
    page: int = 1,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Veritabanından ilanları listele (filtre ve pagination destekli)"""
    listings, total = crud.get_listings(
        db,
        platform=platform,
        kategori=kategori,
        ilan_tipi=ilan_tipi,
        alt_kategori=alt_kategori,
        city=city,
        district=district,
        min_price=min_price,
        max_price=max_price,
        page=page,
        limit=limit
    )

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit,
        "items": [l.to_dict() for l in listings]
    }


@router.get("/listings/{listing_id}")
async def get_listing(listing_id: int, db: Session = Depends(get_db)):
    """Tek bir ilan detayını getir"""
    listing = crud.get_listing_by_id(db, listing_id)
    if not listing:
        raise HTTPException(status_code=404, detail="İlan bulunamadı")
    return listing.to_dict()


@router.get("/sessions")
async def get_scrape_sessions(
    platform: str = None,
    status: str = None,
    page: int = 1,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Tarama oturumlarını listele"""
    sessions, total = crud.get_scrape_sessions(
        db,
        platform=platform,
        status=status,
        page=page,
        limit=limit
    )

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "items": [s.to_dict() for s in sessions]
    }


@router.get("/cities")
async def get_cities(db: Session = Depends(get_db)):
    """Veritabanındaki tüm şehirleri listele"""
    cities = crud.get_all_cities(db)
    return {"cities": cities}


@router.get("/cities/{city}/districts")
async def get_districts(city: str, db: Session = Depends(get_db)):
    """Bir şehrin ilçelerini listele"""
    districts = crud.get_districts_by_city(db, city)
    return {"city": city, "districts": districts}


@router.post("/export/excel")
async def export_to_excel(
    platform: str = None,
    kategori: str = None,
    ilan_tipi: str = None,
    city: str = None,
    district: str = None,
    db: Session = Depends(get_db)
):
    """Filtrelere göre verileri Excel'e aktar ve indir"""
    import pandas as pd
    from datetime import datetime as dt

    # Get listings (max 10000 for export)
    listings, total = crud.get_listings(
        db,
        platform=platform,
        kategori=kategori,
        ilan_tipi=ilan_tipi,
        city=city,
        district=district,
        page=1,
        limit=10000
    )

    if not listings:
        raise HTTPException(status_code=404, detail="Dışa aktarılacak veri bulunamadı")

    # Convert to dataframe
    data = []
    for l in listings:
        row = {
            "Başlık": l.baslik,
            "Fiyat": l.fiyat_text or l.fiyat,
            "Platform": l.platform,
            "Kategori": l.kategori,
            "İlan Tipi": l.ilan_tipi,
            "Alt Kategori": l.alt_kategori,
            "İl": l.location.il if l.location else "",
            "İlçe": l.location.ilce if l.location else "",
            "Mahalle": l.location.mahalle if l.location else "",
            "İlan URL": l.ilan_url,
            "Emlak Ofisi": l.emlak_ofisi,
            "Tarih": l.created_at.strftime("%Y-%m-%d %H:%M") if l.created_at else "",
        }
        # Add details
        if l.details:
            for k, v in l.details.items():
                row[k] = v
        data.append(row)

    df = pd.DataFrame(data)

    # Create temp file
    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    filename = f"export_{timestamp}.xlsx"

    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
    output_dir = os.path.join(project_root, "outputs", "exports")
    os.makedirs(output_dir, exist_ok=True)

    filepath = os.path.join(output_dir, filename)
    df.to_excel(filepath, index=False, engine='openpyxl')

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.delete("/listings/{listing_id}")
async def delete_listing(listing_id: int, db: Session = Depends(get_db)):
    """Bir ilanı sil (soft delete)"""
    listing = crud.get_listing_by_id(db, listing_id)
    if not listing:
        raise HTTPException(status_code=404, detail="İlan bulunamadı")

    listing.is_active = False
    db.commit()

    return {"status": "success", "message": f"İlan {listing_id} silindi"}
